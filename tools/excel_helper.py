import copy
import hashlib
import os
import traceback

from openpyxl.reader.excel import load_workbook
import re
import json
from opencc import OpenCC
from elasticsearch import Elasticsearch
import openpyxl
import jsonlines
import pandas as pd
from datetime import datetime
from openpyxl import Workbook


# :flag: True 简化繁 ; False 繁化简
def shift_name(fullname, flag):
    if flag:
        # 创建 OpenCC 对象，选择转换类型（s2t：简体到繁体，t2s：繁体到简体）
        converter = OpenCC('s2t')
        traditional_chinese = converter.convert(fullname)
        return traditional_chinese
    else:
        # 创建 OpenCC 对象，选择转换类型（s2t：简体到繁体，t2s：繁体到简体）
        converter = OpenCC('t2s')
        traditional_chinese = converter.convert(fullname)
        return traditional_chinese


# 根据excel的sheet 获取所有信息
def dict_reader(sheet):
    results = []
    result = {}
    for ii, row in enumerate(sheet.rows):
        if ii == 0:
            for cell in row:
                val = cell.value
                if val:
                    val = val.strip()
                result[val] = ''
        else:
            deepcopy = copy.deepcopy(result)
            for jj, key in enumerate(deepcopy.keys()):
                if jj < len(row):
                    val = row[jj].value
                    if val is not None:
                        val = str(val).strip()
                    deepcopy[key] = val
            results.append(deepcopy)
    return results


# 得到excel最后一列
def find_last_row(excel_url):
    workbook = openpyxl.load_workbook(excel_url)
    worksheet = workbook['Sheet1']
    last_column_num = int(worksheet.max_column)
    last_column_num = number_to_letter(last_column_num)
    return last_column_num


def before_write_excel(excel_datas):
    excel_save_datas = []
    head_row = list(excel_datas[0].keys())
    excel_save_datas.append(head_row)
    for excel_data in excel_datas:
        excel_data = list(excel_data.values())
        excel_save_datas.append(excel_data)
    return excel_save_datas

# 向excel写入数据
def write_excel(data, excel_url):
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active

    # 写入数据到工作表的多列
    for row, row_data in enumerate(data, start=1):
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=row, column=col, value=value)

    # 保存工作簿
    wb.save(excel_url)

# 根据列数 找到列数对应的字母
def number_to_letter(number):
    if 1 <= number <= 26:
        return chr(ord('A') + number - 1)
    elif number > 26:
        quotient, remainder = divmod(number - 1, 26)
        return number_to_letter(quotient) + number_to_letter(remainder + 1)
    else:
        return None


# 读取excel步骤
def get_excel_data(excel_url):
    workbook = load_workbook(excel_url)
    sheet = workbook[workbook.sheetnames[0]]
    input_records = dict_reader(sheet)
    return input_records


# 获取到链接中的用户id
def get_userid(url):
    user = url.split('/')[-1]
    if len(user) == 0:
        user = url.split('/')[-2]
    start_index = user.find('id=')
    if start_index != -1:
        user = user[start_index + len("id="):]
    return user


# 导出 jsonlines 文件 datas存储列表形式的jsonlines [{},{}]
def output_jsonlines(url, datas):
    #    新增 去重 后处理
    datas = move_repetition(datas)
    try:
        with open(url, 'w', encoding='utf-8', buffering=2048) as f:
            batch_size = 100
            for i in range(0, len(datas), batch_size):
                batch = datas[i:i+batch_size]
                f.writelines(str(data) + "\n" for data in batch)
    except Exception as e:
        print(e)


# 进行md5运算
def get_md5(value):
    # 创建一个md5对象
    md5 = hashlib.md5()
    md5.update(value.encode('utf-8'))
    result_md5 = md5.hexdigest()
    return result_md5


# 拼接单个object对象的json
# flag 为True时追加到同一个字典中
def joint_object_json(json_row, head_key, key, value, flag, dic):
    if flag:
        if head_key not in json_row:
            json_row[head_key] = [{}]
        json_row[head_key][0][key] = value
    else:
        if head_key not in json_row:
            json_row[head_key] = []
        dic_copy = copy.deepcopy(dic)
        json_row[head_key].append(dic_copy)


# 1.去除特殊字符及空格
# 2.首位为0去除
# 3.如果开头不是886 进行添加
def deal_phonenum(number):
    # 去除特殊字符及空格 保留字母
    number = re.sub(r'\W+', '', number)
    # 去除首位的0
    number = number.lstrip('0')
    # 如果开头不是886，则添加在开头
    if not number.startswith('886'):
        number = '886' + number
    return number


# 金额带有逗号的字符串转为int
def account_string_to_int(account_string):
    if type(account_string) == int:
        return account_string
    if ',' in account_string:
        value_without_comma = account_string.replace(',', '')
        float_value = float(value_without_comma)
        result = int(float_value)
        return result


# 对存储年份的字符串进行处理
# 1. 去除特殊字符
# 2.判断八位则前四位未开始时间，后四位为结束时间
# 3. 只有四位为开始时间
def deal_crimes_year(year):
    year = re.sub(r'\D', '', year)
    if len(year) == 8:
        return year[:4], year[4:]
    return year[:4], None


# 处理日期格式
def deal_date(date):
    try:
        date = re.sub(r'\D+', '', date)
        if len(date) < 6:
            return ''
        date = date[:8]
        # 解析日期字符串
        date = datetime.strptime(date, "%Y%m%d")

        # 格式化为 YYYY-MM-DD 格式
        formatted_date = date.strftime("%Y-%m-%d")
        return formatted_date
    except Exception as e:
        print(e)


# 处理日期格式 2022-04-19 16:11:01  -----  2024年01月01日
def deal_date_chinese(date):
    if not date:
        return ''
    # 将原始日期字符串转换为datetime对象
    date_obj = datetime.strptime(date, '%Y-%m-%d %H:%M:%S')

    # 将datetime对象格式化为指定格式的字符串
    formatted_date = date_obj.strftime('%Y年%m月%d日')
    return formatted_date


# 处理日期格式 2024年01月01日 ----- 2022-04-19
def deal_date_chinese2symbol(date):
    if not date:
        return ''
    # 将原始日期字符串转换为datetime对象
    date_obj = datetime.strptime(date, '%Y年%m月%d日')

    # 将datetime对象格式化为指定格式的字符串
    formatted_date = date_obj.strftime('%Y-%m-%d')
    return formatted_date


# 处理日期格式 时间戳  -----  2013-09-26 22:02:44
def deal_date_timestamp(timestamp):
    if not timestamp:
        return ''
    if type(timestamp) == str:
        timestamp = int(timestamp)
    dt = datetime.fromtimestamp(timestamp)
    formatted_time = dt.strftime("%Y-%m-%d %H:%M:%S")
    return formatted_time


# 对可能存在一个单元格多个email数据进行处理
# 返回值是 一个列表 直接进行key value替换
def deal_email(email):
    email_json = []
    if email.count('@') > 1:
        emails = email.split(';')
        for email_one in emails:
            dic = {'value': email_one, 'domain': email_one.split('@')[-1]}
            email_json.append(dic)
    else:
        dic = {'value': email, 'domain': email.split('@')[-1]}
        email_json.append(dic)
    return email_json


# 截取邮箱域名
def split_domain_name(email):
    return email.split("@")[-1]


# 处理中文名称
def deal_chinese_name(value):
    if type(value) is not str:
        return value
    chinese_pattern = re.compile(r'[^\u4e00-\u9fa5·]')
    # 查找第一个特殊字符的位置
    match = chinese_pattern.search(value)
    if match is not None:
        return value[:match.start()]

    return value


# 替换英文符号
def repeat_symbol_string(data):
    data = data.replace("，", ",").replace("。", ".").replace("！", "!").replace("；", ";").replace("【", "[")\
        .replace("】", "]").replace("（", "(").replace("）", ")").replace("{", "{").replace("}", "}").replace("：", ":")
    return data


# 分隔长字符串 (12),(23)-> (12)(23)
def split_string(data, split_str):
    data = data.replace("，", ",").replace("。", ".").replace("！", "!").replace("；", ";").replace("【", "[")\
        .replace("】", "]").replace("（", "(").replace("）", ")").replace("{", "{").replace("}", "}")
    result = []
    if split_str in data:
        result = data.split(split_str)
    result.append(data)
    return result


# 根据) 分隔字符串 (12)(23)-> (12)(23)
def split_last_string(data, split_str):
    data = data.replace("，", ",").replace("。", ".").replace("！", "!").replace("；", ";").replace("【", "[")\
        .replace("】", "]").replace("（", "(").replace("）", ")").replace("{", "{").replace("}", "}")
    pattern = r'(?<=' + split_str + ')'
    if split_str is data:
        data = re.split(pattern, data)
    return data


# 读取jsonline数据并 返回所有jsonline数据
def read_jsonline(jsonline_url):
    all_jsonlines = []
    with open(jsonline_url, 'r', encoding='utf-8') as f:
        for line in f:
            li = json.loads(line)
            all_jsonlines.append(li)
    return all_jsonlines


# 读取txt文件内容
def read_txt(txt_url):
    with open(txt_url, 'r', encoding='utf-8') as file:
        content = file.read()
        return content


# :all_jsonlines: 所有的jsonline数据
# :number: 需要匹配的数据
# ：return：返回符合数据json列表
def get_number_jsonline(all_jsonlines, number):
    result = []
    for index, jsonline in enumerate(all_jsonlines):
        if number in jsonline:
            jsondata = jsonline[number]
            result.append(jsondata)
    return result


# 对jsonline进行去重
def move_repetition(jsonlines):
    try:
        result = []
        for jsonline in jsonlines:
            for key, value in jsonline.items():
                if type(jsonline[key]) == list and len(jsonline[key]) != 0:
                    jsonline[key] = distinct_list_string_dict(jsonline[key])
            # jsonlines 修改录入数据为字符串形式
            jsonline = json.dumps(jsonline, ensure_ascii=False)
            result.append(jsonline)
        return result
    except Exception as e:
        print(e)


# 对数据 进行去重
def move_repetition_data(datas):
    try:
        result = []
        for data in datas:
            for key, value in data.items():
                if type(data[key]) == list and len(data[key]) != 0:
                    data[key] = distinct_list_string_dict(data[key])
            result.append(data)
        return result
    except Exception as e:
        print(e)


# 对字典数组进行去重
def distinct_list_dict(docs):
    merge_doc_dict = {}
    for doc in docs:
        new_doc = {}

        for key_tmp in sorted(doc):
            new_doc[key_tmp] = json.dumps(doc[key_tmp])

        merge_doc_dict[get_md5(json.dumps(new_doc))] = doc

    if merge_doc_dict:
        return list(merge_doc_dict.values())


# 对字典数组 和 字符串数组 进行去重
def distinct_list_string_dict(docs):
    merge_doc_dict = {}
    for doc in docs:
        new_doc = {}
        if type(doc) == dict:
            flag = True
            for key_tmp in sorted(doc):
                new_doc[key_tmp] = json.dumps(doc[key_tmp])
            merge_doc_dict[get_md5(json.dumps(new_doc))] = doc
        else:
            flag = False
            merge_doc_dict[doc] = None

    if flag:
        if merge_doc_dict:
            return list(merge_doc_dict.values())
    else:
        return list(set(merge_doc_dict.keys()))


# 处理jsonlines中 personId 重复的值
def deal_jsonlines_remove_duplicates(input_file, output_file):
    unique_records = {}

    with jsonlines.open(input_file, 'r') as reader:
        for record in reader:
            person_id = record.get('personId')
            if person_id not in unique_records:
                unique_records[person_id] = record
        reader.close()

    with jsonlines.open(output_file, 'w') as writer:
        for _, record in unique_records.items():
            writer.write(record)
        writer.close()


# 判断是否为台湾身份证号码
def is_taiwan_identity(number):
    if number:
        match = re.match(r'^[a-zA-Z][0-9]{9}$', number)
        return match


# 字符串 去除remove_str的内容
def remove_fixed_chinese_characters(text, remove_str):
    pattern = '[' + remove_str + ']'
    result = re.sub(pattern, '', text)
    return result


# 返回第一个数字之前的内容
def get_prefix_before_first_digit_string(string):
    match = re.search(r'^\D+', string)
    if match:
        return string[:match.end()]
    else:
        return string


# 根据1.2.3.截取内容 截取结果：1.…… 2.…… 3.……
def split_findall_number(text):
    # 使用正则表达式匹配包含数字序号的文本块
    matches = re.findall(r'\d+\. .*?(?=\n\d|\n*$)', text, re.DOTALL)
    return matches
    # 遍历每个文本块并输出
    # for i, sub_text in enumerate(matches, start=1):
    #     print(f"{i}. {sub_text.strip()}")


# 根据010203截取内容 截取结果：01…… 02…… 03……
def split_findall_number_zero(text):

    # 正则表达式模式
    pattern = r"\b0\d\b.*?(?=\b\d{2}\b|$)"


    # 使用 re.findall() 方法匹配所有符合模式的子串，并使用 strip() 方法去除空格
    matches = re.findall(pattern, text, re.DOTALL)
    return matches


def jsonlines_to_excel(jsonlines_url, excel_url):
    # 读取 JSON Lines 文件
    df = pd.read_json(jsonlines_url, lines=True)
    # 将数据保存为 Excel 文件
    df.to_excel(excel_url, index=False)


# 判断文件 是否存在，不存在创建路径及其文件
def create_file_if_not_exists(file_path):
    # 判断文件路径是否存在
    if not os.path.exists(file_path):
        # 创建文件目录
        os.makedirs(os.path.dirname(file_path), exist_ok=True)

        # 创建空文件
        with open(file_path, 'w') as file:
            pass


query = {
    "query": {
        "bool": {
            "should": []
        }
    }
}


# 查询 es
def query_es(ip_addr, query_template, index_name):
    es = Elasticsearch(ip_addr)
    result = es.search(index=index_name, body=query_template)
    return result


# 拼接es查询语句
def joint_es_query(cell_values):
    query_template = copy.deepcopy(query)
    should_list = []
    for row_value in cell_values:
        should_list.append(joint_es_conditions(row_value))
    query_template['query']['bool']['should'].append(should_list)
    return query_template